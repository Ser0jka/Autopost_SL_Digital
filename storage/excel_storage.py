import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from models.post import Post

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / "data" / "posts.xlsx"
SHEET_NAME = "posts"

COLUMNS = [
    "id", "created_at", "updated_at", "rubric_id", "topic", "caption",
    "image_prompt", "image_path", "image_provider", "status",
    "approved_by", "published_at", "telegram_message_id", "error_message",
]


class ExcelStorage:
    def __init__(self, path: Path = EXCEL_PATH):
        self.path = path
        self._ensure_file()

    def _ensure_file(self) -> None:
        if not self.path.exists():
            self.path.parent.mkdir(parents=True, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append(COLUMNS)
            ws.freeze_panes = "A2"
            widths = [25, 20, 20, 15, 50, 80, 80, 50, 15, 15, 15, 20, 20, 50]
            for idx, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = w
            wb.save(self.path)
            logger.info(f"Created Excel: {self.path}")
            return

        wb = openpyxl.load_workbook(self.path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        ws.title = SHEET_NAME
        existing = [cell.value for cell in ws[1]]
        normalized = ["caption" if value == "text" else value for value in existing]
        keep_columns = [column for column in normalized if column in COLUMNS]

        if keep_columns == COLUMNS:
            return

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(normalized, row))
            if row_data.get("id"):
                rows.append([row_data.get(column) for column in COLUMNS])

        old_ws = ws
        ws = wb.create_sheet(f"{SHEET_NAME}_new", 0)
        ws.append(COLUMNS)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        widths = [25, 20, 20, 15, 50, 80, 80, 50, 15, 15, 15, 20, 20, 50]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        wb.remove(old_ws)
        ws.title = SHEET_NAME
        wb.save(self.path)
        logger.info("Migrated Excel schema: %s", self.path)

    def _load(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.path)

    def _save(self, wb: openpyxl.Workbook) -> None:
        wb.save(self.path)

    def _col_index(self) -> dict:
        return {name: idx for idx, name in enumerate(COLUMNS)}

    def _row_to_post(self, row) -> Post:
        vals = [cell.value for cell in row]
        data = dict(zip(COLUMNS, vals))

        def _dt(v) -> Optional[datetime]:
            return v if isinstance(v, datetime) else None

        def _s(v) -> str:
            return str(v) if v is not None else ""

        def _i(v) -> Optional[int]:
            try:
                return int(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        return Post(
            id=_s(data.get("id")),
            created_at=_dt(data.get("created_at")) or datetime.now(),
            updated_at=_dt(data.get("updated_at")) or datetime.now(),
            rubric_id=_s(data.get("rubric_id")),
            topic=_s(data.get("topic")),
            caption=_s(data.get("caption")),
            image_prompt=_s(data.get("image_prompt")),
            image_path=_s(data.get("image_path")),
            image_provider=_s(data.get("image_provider")),
            status=_s(data.get("status")) or "draft",
            approved_by=_i(data.get("approved_by")),
            published_at=_dt(data.get("published_at")),
            telegram_message_id=_i(data.get("telegram_message_id")),
            error_message=_s(data.get("error_message")) or None,
        )

    def add_post(self, post: Post) -> None:
        wb = self._load()
        ws = wb[SHEET_NAME]
        ws.append([
            post.id, post.created_at, post.updated_at, post.rubric_id,
            post.topic, post.caption, post.image_prompt, post.image_path,
            post.image_provider, post.status, post.approved_by,
            post.published_at, post.telegram_message_id, post.error_message,
        ])
        self._save(wb)
        logger.info(f"Added post {post.id}")

    def update_post(self, post_id: str, fields: dict) -> bool:
        wb = self._load()
        ws = wb[SHEET_NAME]
        col_idx = self._col_index()

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == post_id:
                for key, value in fields.items():
                    if key in col_idx:
                        row[col_idx[key]].value = value
                if "updated_at" not in fields:
                    row[col_idx["updated_at"]].value = datetime.now()
                self._save(wb)
                logger.info(f"Updated post {post_id}: {list(fields.keys())}")
                return True

        logger.warning(f"Post {post_id} not found for update")
        return False

    def get_post(self, post_id: str) -> Optional[Post]:
        wb = self._load()
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=2):
            if row[0].value and str(row[0].value) == post_id:
                return self._row_to_post(row)
        return None

    def list_recent_posts(self, limit: int = 20) -> List[Post]:
        wb = self._load()
        ws = wb[SHEET_NAME]
        posts = [self._row_to_post(row) for row in ws.iter_rows(min_row=2) if row[0].value]
        return posts[-limit:]

    def list_posts_by_status(self, status: str) -> List[Post]:
        wb = self._load()
        ws = wb[SHEET_NAME]
        col_idx = self._col_index()
        sc = col_idx["status"]
        return [
            self._row_to_post(row)
            for row in ws.iter_rows(min_row=2)
            if row[0].value and str(row[sc].value) == status
        ]

    def export_path(self) -> Path:
        return self.path
